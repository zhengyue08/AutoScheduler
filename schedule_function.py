import random
import pandas as pd
import openpyxl
import csv

# unchangable values for every week and for any task

days = ["星期一", "星期二", "星期三", "星期四", "星期五",]

periods = ["12节", "34节", "56节", "78节"]

shifts = [day + " " + period for day in days for period in periods]


# load names firstly
def get_names():
    with open('names.csv', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            return row
        
def write_names_file(names):
    # Open the CSV file for writing
    with open('names.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        # Write the data to the CSV file
        writer.writerow(names)

# NOTE: 2nd task
def add_member(name, names=get_names()):
    names.append(name)
    write_names_file(names)

# NOTE: 3rd task
def delete_member(name, names=get_names()):
    names.remove(name)
    write_names_file(names)


# read unavailable excel to available dictionary
def unavailable_excel_to_availabe_dictionary(excel_name_path, names=get_names()):
    unavailable_df = pd.read_excel(excel_name_path).dropna(axis=0, how='all')

    unavailable = {}
    available = {}

    # store unavailable periods in a dictionary unavailable = e.g. "星期一 12节":["张三", "李四"]
    for _, row in unavailable_df.iterrows():
        name, day, period = row
        # print(name, shift, unavailabilities)
        shift = str(day) + " " + period
        unavailable.setdefault(shift,[]).append(name)
    # store available periods in a dictionary: available = e.g. "星期一 12节":["张三", "李四"]
    for shift in unavailable:
        available[shift] = [name for name in names if name not in unavailable[shift]]   

    return available

# read available excel to available dictionary
def available_excel_to_available_dictionary(excel_name_path):

    available_df = pd.read_excel(excel_name_path).dropna(axis=0, how='all')

    available = {}

    # store available periods in a dictionary: available = e.g. "星期一 12节":["张三", "李四"]
    for _, row in available_df.iterrows():
        name, day, period = row
        print(name, day, period)
        shift = str(day) + " " + str(period)
        available.setdefault(shift,[]).append(name)   

    return available


def generate_schedule(week, names):
    # generate schedule for week
    # read unavailable excel to available dictionary
    available = unavailable_excel_to_availabe_dictionary("available_" + str(week) + ".xlsx")
    

    # define two lists to store the number of shifts assigned for each person and the shifts of each person
    person_shift_counts = {name: 0 for name in names}
    person_shift_times = {name: [] for name in names}

    # print(person_shift_counts)
    # print(person_shift_times)

    # high_priority -> two people per shift
    high_priority_shifts = ["星期二 34节", "星期二 56节", "星期四 34节", "星期四 56节"]
    normal_shifts = [shift for shift in shifts if shift not in high_priority_shifts]
    print(normal_shifts)

    # store assigned shifts 
    assigned_by_person = {}
    assigned_by_shift = {}

    # Prioritize high-priority shifts for assignment
    for shift in high_priority_shifts:
        unassigned = names.copy()

        # assign two people per shift
        for i in range(2):
            # get available people 
            available_for_this_shift = [name for name in unassigned if person_shift_counts[name] < 2 and shift not in person_shift_times[name] and name in available.get((shift), [])]
            if len(available_for_this_shift) == 0:
                print("No available person for shift: ", shift)
                continue
            # assign person randomly
            name = random.choice(available_for_this_shift)

            # update available according to conditions
            person_shift_counts[name] += 1
            person_shift_times[name].append(shift)

            assigned_by_person.setdefault(name,[]).append(shift)
            assigned_by_shift.setdefault(shift,[]).append(name)

            # avoid assigning the same person to the same shift twice
            unassigned.remove(name)

    # for shift in high_priority_shifts:
    #     print(shift, assigned_by_shift[shift])
    #     print(shift, available[shift])

    # assign remaining shifts: make sure that every shift has at least one person assigned
    for shift in normal_shifts:
        unassigned = names.copy()
        for i in range(1):
            # get available people 
            available_for_this_shift = [name for name in unassigned if person_shift_counts[name] < 2 and shift not in person_shift_times[name] and name in available.get((shift), [])]
            if len(available_for_this_shift) == 0:
                print("No available person for shift under the condition that each person is assigned to at most two shifts, try to assign someone for three shifts in this week", shift)
                available_for_this_shift = [name for name in unassigned if person_shift_counts[name] < 3 and shift not in person_shift_times[name] and name in available.get((shift), [])]

            # assign person randomly
            name = random.choice(available_for_this_shift)

            # update available according to conditions
            person_shift_counts[name] += 1
            person_shift_times[name].append(shift)

            assigned_by_person.setdefault(name,[]).append(shift)
            assigned_by_shift.setdefault(shift,[]).append(name)

            # avoid assigning the same person to the same shift twice
            unassigned.remove(name)

    # assign one more shift to person who has only one shift assigned
    one_shift_person = [name for name in names if person_shift_counts[name] == 1 or person_shift_counts[name] == 0]
    one_person_shift = [shift for shift in shifts if len(assigned_by_shift[shift]) == 1]

    for name in one_shift_person:
        for shift in one_person_shift:
            if name in available.get((shift), []):
                person_shift_counts[name] += 1
                person_shift_times[name].append(shift)

                assigned_by_person.setdefault(name,[]).append(shift)
                assigned_by_shift.setdefault(shift,[]).append(name)

                # avoid assigning the same person to the same shift twice
                one_person_shift.remove(shift)
                break

    print(len(assigned_by_shift))
    print(assigned_by_shift)
    print(person_shift_counts)

    # draw final schedule file
    draw_final_schedule_file(assigned_by_shift, week, person_shift_counts, names)


def draw_final_schedule_file(assigned_by_shift, week, person_shift_counts, names):

    # 创建一个Workbook对象
    wb = openpyxl.load_workbook('temple.xlsx')

    # 获取活跃的工作表对象
    ws = wb.active

    # 写入表头
    ws['A1'] = "第" + str(week) + "周值班表"


    ws['B3'] = ' '.join(assigned_by_shift['星期一 12节'])
    ws['B4'] = ' '.join(assigned_by_shift['星期一 34节'])
    ws['B5'] = ' '.join(assigned_by_shift['星期一 56节'])
    ws['B6'] = ' '.join(assigned_by_shift['星期一 78节'])


    ws['C3'] = ' '.join(assigned_by_shift['星期二 12节'])
    ws['C4'] = ' '.join(assigned_by_shift['星期二 34节'])
    ws['C5'] = ' '.join(assigned_by_shift['星期二 56节'])
    ws['C6'] = ' '.join(assigned_by_shift['星期二 78节'])

    ws['D3'] = ' '.join(assigned_by_shift['星期三 12节'])
    ws['D4'] = ' '.join(assigned_by_shift['星期三 34节'])
    ws['D5'] = ' '.join(assigned_by_shift['星期三 56节'])
    ws['D6'] = ' '.join(assigned_by_shift['星期三 78节'])

    ws['E3'] = ' '.join(assigned_by_shift['星期四 12节'])
    ws['E4'] = ' '.join(assigned_by_shift['星期四 34节'])
    ws['E5'] = ' '.join(assigned_by_shift['星期四 56节'])
    ws['E6'] = ' '.join(assigned_by_shift['星期四 78节'])

    ws['F3'] = ' '.join(assigned_by_shift['星期五 12节'])
    ws['F4'] = ' '.join(assigned_by_shift['星期五 34节'])
    ws['F5'] = ' '.join(assigned_by_shift['星期五 56节'])
    ws['F6'] = ' '.join(assigned_by_shift['星期五 78节'])

    for cell in range(8,8+len(names)):
        ws["D"+str(cell)] = names[cell-8]
        ws["E"+str(cell)] = person_shift_counts[names[cell-8]]

    wb.save("第" + str(week) + "周值班表最终版.xlsx")


def generate_temple_available(week, names=get_names()):
# TODO: 4th task
    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    for name in names:
        for day in days:
            for period in periods:
                worksheet.append([name, day, period])

    workbook.save("available_" + str(week) + ".xlsx")