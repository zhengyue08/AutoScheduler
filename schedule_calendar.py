import pandas as pd
from openpyxl import Workbook


# 创建一个Workbook对象
wb = Workbook()

# 获取活跃的工作表对象
ws = wb.active

# 写入表头
ws.append(['星期一', '星期二', '星期三', '星期四', '星期五'])

# 读取schedule
schedule_df = pd.read_excel('schedule.xlsx')
print(schedule_df)

# 转换schedule_df为schedule字典
schedule = {'星期一 12节': []}

# a = schedule['星期一 12节']
# a.append('lisa')

# print(a)


for _, row in schedule_df.iterrows():
    name, shift = row
    schedule.setdefault(shift,[]).append(name)

print(schedule)
# print(schedule)

ws['A1'] = ""
ws['A2'] = "12节"
ws['A3'] = "34节"
ws['A4'] = "56节"
ws['A5'] = "78节"

ws['B1'] = "星期一"
ws['C1'] = "星期二"
ws['D1'] = "星期三"
ws['E1'] = "星期四"
ws['F1'] = "星期五"

ws['B2'] = ' '.join(schedule['星期一 12节'])
ws['B3'] = ' '.join(schedule['星期一 34节'])
ws['B4'] = ' '.join(schedule['星期一 56节'])
ws['B5'] = ' '.join(schedule['星期一 78节'])


ws['C2'] = ' '.join(schedule['星期二 12节'])
ws['C3'] = ' '.join(schedule['星期二 34节'])
ws['C4'] = ' '.join(schedule['星期二 56节'])
ws['C5'] = ' '.join(schedule['星期二 78节'])

ws['D2'] = ' '.join(schedule['星期三 12节'])
ws['D3'] = ' '.join(schedule['星期三 34节'])
ws['D4'] = ' '.join(schedule['星期三 56节'])
ws['D5'] = ' '.join(schedule['星期三 78节'])

ws['E2'] = ' '.join(schedule['星期四 12节'])
ws['E3'] = ' '.join(schedule['星期四 34节'])
ws['E4'] = ' '.join(schedule['星期四 56节'])
ws['E5'] = ' '.join(schedule['星期四 78节'])

ws['F2'] = ' '.join(schedule['星期五 12节'])
ws['F3'] = ' '.join(schedule['星期五 34节'])
ws['F4'] = ' '.join(schedule['星期五 56节'])
ws['F5'] = ' '.join(schedule['星期五 78节'])

wb.save('schedule_calendarview.xlsx')











